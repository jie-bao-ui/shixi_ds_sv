import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# ====== 改成你的两个文件路径 ======
FILE_A ="D:\\uid03859\\Desktop\\python\\table_1_15_03_32_35_to_03_35_21.txt"
FILE_B ="D:\\uid03859\\Desktop\\python\\table_1_16_03_32_35_to_03_35_21.txt"
# =================================

OUT_CSV  = "summary_stats.csv"
OUT_XLSX = "summary_stats.xlsx"

DISTANCES = [0, 10, 20, 30, 40]

def parse_pipe_list(s: str, cast=float):
    s = str(s).strip()
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1]
    if s == "":
        return []
    return [cast(x) for x in s.split("|")]

def load_long_bias(file_path: str) -> pd.DataFrame:
    """
    读入并展开为 long:
      timestamp(str), distance(int), bias(float)
    规则：
      1) 同一 timestamp 的重复行保留第一个
      2) 只保留 valid==1 的 bias
      3) 展开后同一 (timestamp, distance) 重复保留第一个
    """
    df = pd.read_csv(file_path)

    # 1) 同一 timestamp 重复保留第一个（按文件顺序）
    df = df.drop_duplicates(subset=["timestamp"], keep="first").reset_index(drop=True)

    df["dist_list"]  = df["distances"].apply(lambda x: parse_pipe_list(x, int))
    df["valid_list"] = df["validty"].apply(lambda x: parse_pipe_list(x, int))
    df["bias_list"]  = df["bias"].apply(lambda x: parse_pipe_list(x, float))

    records = []
    for row_no, (ts, dists, valids, biases) in enumerate(
        zip(df["timestamp"], df["dist_list"], df["valid_list"], df["bias_list"])
    ):
        for d, v, b in zip(dists, valids, biases):
            if v == 1:
                records.append({
                    "row_no": row_no,          # 用于“保留第一个”
                    "timestamp": ts,
                    "distance": int(d),
                    "bias": float(b),
                })

    long_df = pd.DataFrame(records)
    if long_df.empty:
        return long_df

    # 3) 展开后：同一 (timestamp, distance) 重复保留第一个
    long_df = (long_df
               .sort_values(["timestamp", "row_no"])
               .drop_duplicates(subset=["timestamp", "distance"], keep="first")
               .reset_index(drop=True))

    return long_df

def stats_by_distance(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    输出 index=distance 的统计表：mean/std
    """
    if long_df.empty:
        return pd.DataFrame(index=DISTANCES, data={"mean": [pd.NA]*len(DISTANCES),
                                                   "std":  [pd.NA]*len(DISTANCES)})

    g = long_df.groupby("distance")["bias"]
    out = pd.DataFrame({
        "mean": g.mean(),
        "std":  g.std(ddof=1)  # 样本标准差
    })

    # 补全缺失距离行
    out = out.reindex(DISTANCES)
    return out

# 读取并统计
A_long = load_long_bias(FILE_A)
B_long = load_long_bias(FILE_B)

A_stats = stats_by_distance(A_long).rename(columns={"mean": "mean_docA", "std": "std_docA"})
B_stats = stats_by_distance(B_long).rename(columns={"mean": "mean_docB", "std": "std_docB"})

summary = pd.concat([A_stats, B_stats], axis=1)
summary.index.name = "distance"
summary = summary.reset_index()

# 计算“绝对值较大者”标记（CSV 用 * 标记；XLSX 用红色字体）
def bigger_abs(a, b):
    if pd.isna(a) and pd.isna(b):
        return "tie"
    if pd.isna(a):
        return "docB"
    if pd.isna(b):
        return "docA"
    return "docA" if abs(a) > abs(b) else ("docB" if abs(b) > abs(a) else "tie")

summary["mean_abs_bigger"] = summary.apply(lambda r: bigger_abs(r["mean_docA"], r["mean_docB"]), axis=1)
summary["std_abs_bigger"]  = summary.apply(lambda r: bigger_abs(r["std_docA"],  r["std_docB"]),  axis=1)

# 给 CSV 增加“星号标记”列（CSV 本身不支持红色字体）
summary["mean_docA_mark"] = summary.apply(lambda r: "*" if r["mean_abs_bigger"] == "docA" else "", axis=1)
summary["mean_docB_mark"] = summary.apply(lambda r: "*" if r["mean_abs_bigger"] == "docB" else "", axis=1)
summary["std_docA_mark"]  = summary.apply(lambda r: "*" if r["std_abs_bigger"]  == "docA" else "", axis=1)
summary["std_docB_mark"]  = summary.apply(lambda r: "*" if r["std_abs_bigger"]  == "docB" else "", axis=1)

# 输出 CSV（带标记）
summary.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

# 输出 XLSX（真正“标红”）
wb = Workbook()
ws = wb.active
ws.title = "summary"

headers = list(summary.columns)
ws.append(headers)

red_font = Font(color="FF0000")

for i, row in enumerate(summary.itertuples(index=False), start=2):
    values = list(row)
    ws.append(values)

    # 找到对应列号
    col_mean_A = headers.index("mean_docA") + 1
    col_mean_B = headers.index("mean_docB") + 1
    col_std_A  = headers.index("std_docA")  + 1
    col_std_B  = headers.index("std_docB")  + 1
    col_mean_flag = headers.index("mean_abs_bigger") + 1
    col_std_flag  = headers.index("std_abs_bigger")  + 1

    mean_flag = ws.cell(i, col_mean_flag).value
    std_flag  = ws.cell(i, col_std_flag).value

    if mean_flag == "docA":
        ws.cell(i, col_mean_A).font = red_font
    elif mean_flag == "docB":
        ws.cell(i, col_mean_B).font = red_font

    if std_flag == "docA":
        ws.cell(i, col_std_A).font = red_font
    elif std_flag == "docB":
        ws.cell(i, col_std_B).font = red_font

wb.save(OUT_XLSX)

print(f"[OK] CSV  已生成: {OUT_CSV}")
print(f"[OK] XLSX 已生成(红色标注): {OUT_XLSX}")