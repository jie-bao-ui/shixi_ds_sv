import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# ====== 改成你的两个文件路径 ======
FILE_A = "D:\\uid03859\\Desktop\\python\\table_1_15_03_32_35_to_03_35_21.txt"
FILE_B = "D:\\uid03859\\Desktop\\python\\table_1_16_03_32_35_to_03_35_21.txt"
# =================================

OUT_XLSX = "exceed_stats_by_distance.xlsx"

# 距离与阈值映射：0/10/20/30/40 对应 9/9/11/14/16
THRESH = {0: 9, 10: 9, 20: 11, 30: 14, 40: 16}
DISTANCES = [0, 10, 20, 30, 40]

def parse_pipe_list(s: str, cast=float):
    s = str(s).strip()
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1]
    if s == "":
        return []
    return [cast(x) for x in s.split("|")]

def load_long_bias(file_path: str) -> pd.DataFrame:
    """
    读入并展开成 long 表：timestamp, distance, bias
    规则：
      - 同一 timestamp 的重复行保留第一个（文件顺序）
      - 只保留 validty==1 的点
      - 展开后同一 (timestamp, distance) 重复保留第一个
    """
    df = pd.read_csv(file_path)

    # timestamp 重复保留第一个（按文件顺序）
    df = df.drop_duplicates(subset=["timestamp"], keep="first").reset_index(drop=True)

    df["dist_list"]  = df["distances"].apply(lambda x: parse_pipe_list(x, int))
    df["valid_list"] = df["validty"].apply(lambda x: parse_pipe_list(x, int))
    df["bias_list"]  = df["bias"].apply(lambda x: parse_pipe_list(x, float))

    records = []
    for row_no, (ts, dists, valids, biases) in enumerate(
        zip(df["timestamp"], df["dist_list"], df["valid_list"], df["bias_list"])
    ):
        for d, v, b in zip(dists, valids, biases):
            if v == 1:
                records.append({
                    "row_no": row_no,
                    "timestamp": ts,
                    "distance": int(d),
                    "bias": float(b),
                })

    long_df = pd.DataFrame(records)
    if long_df.empty:
        return long_df

    # 展开后：同一 (timestamp, distance) 重复保留第一个
    long_df = (long_df
               .sort_values(["timestamp", "row_no"])
               .drop_duplicates(subset=["timestamp", "distance"], keep="first")
               .reset_index(drop=True))
    return long_df

def compute_stats(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    输出每个 distance 的 total_valid_cnt / exceed_cnt / exceed_pct
    exceed 条件：abs(bias) > THRESH[distance]
    """
    rows = []
    for d in DISTANCES:
        sub = long_df[long_df["distance"] == d]
        total = int(len(sub))
        if total == 0:
            exceed = 0
            pct = None
        else:
            exceed = int((sub["bias"].abs() > THRESH[d]).sum())
            pct = exceed / total
        rows.append({
            "distance": d,
            "threshold": THRESH[d],
            "total_valid_cnt": total,
            "exceed_cnt": exceed,
            "exceed_pct": pct
        })
    return pd.DataFrame(rows)

# 读取并计算
A_long = load_long_bias(FILE_A)
B_long = load_long_bias(FILE_B)

A_stats = compute_stats(A_long).rename(columns={
    "total_valid_cnt": "total_valid_cnt_A",
    "exceed_cnt": "exceed_cnt_A",
    "exceed_pct": "exceed_pct_A",
})
B_stats = compute_stats(B_long).rename(columns={
    "total_valid_cnt": "total_valid_cnt_B",
    "exceed_cnt": "exceed_cnt_B",
    "exceed_pct": "exceed_pct_B",
})

# 合并（distance/threshold 相同）
out = pd.merge(
    A_stats[["distance","threshold","total_valid_cnt_A","exceed_cnt_A","exceed_pct_A"]],
    B_stats[["distance","threshold","total_valid_cnt_B","exceed_cnt_B","exceed_pct_B"]],
    on=["distance","threshold"],
    how="outer"
).sort_values("distance").reset_index(drop=True)

# 写 xlsx 并标红：对每一行，exceed_pct 绝对值更大（这里 pct 本身非负）者标红
wb = Workbook()
ws = wb.active
ws.title = "exceed_stats"

headers = [
    "distance", "threshold",
    "total_valid_cnt_A", "exceed_cnt_A", "exceed_pct_A",
    "total_valid_cnt_B", "exceed_cnt_B", "exceed_pct_B",
]
ws.append(headers)

red_font = Font(color="FF0000")

def fmt_pct(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    return float(x)

for r_idx, row in enumerate(out.itertuples(index=False), start=2):
    # 写一行
    values = [
        row.distance, row.threshold,
        row.total_valid_cnt_A, row.exceed_cnt_A, fmt_pct(row.exceed_pct_A),
        row.total_valid_cnt_B, row.exceed_cnt_B, fmt_pct(row.exceed_pct_B),
    ]
    ws.append(values)

    # 百分比列索引（1-based）
    col_pct_A = headers.index("exceed_pct_A") + 1
    col_pct_B = headers.index("exceed_pct_B") + 1

    a = row.exceed_pct_A
    b = row.exceed_pct_B

    # 只有两边都有值才比较
    if a is not None and not (isinstance(a, float) and pd.isna(a)) and \
       b is not None and not (isinstance(b, float) and pd.isna(b)):
        if a > b:
            ws.cell(r_idx, col_pct_A).font = red_font
        elif b > a:
            ws.cell(r_idx, col_pct_B).font = red_font
        else:
            # 相等不标红（如需都标红可改）
            pass

# 把百分比列设置为百分比显示
for row in ws.iter_rows(min_row=2, min_col=col_pct_A, max_col=col_pct_B):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"

wb.save(OUT_XLSX)
print(f"[OK] 已生成: {OUT_XLSX}")